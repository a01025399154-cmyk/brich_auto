# -*- coding: utf-8 -*-
import gspread
from google.oauth2.service_account import Credentials
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os
import shutil
from datetime import datetime
import openpyxl
import requests
from bs4 import BeautifulSoup
import signal
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
import config  # 설정 파일 임포트

# 전역 변수로 중단 신호 관리
interrupted = False

def signal_handler(sig, frame):
    """Ctrl+C 신호를 처리하는 핸들러"""
    global interrupted
    print('\n\n⚠️ 중단 신호를 받았습니다. 안전하게 종료합니다...')
    interrupted = True
    # 모든 프로세스 강제 종료
    try:
        # 현재 실행 중인 Chrome 프로세스 종료
        os.system("pkill -f chrome")
        os.system("pkill -f chromedriver")
    except:
        pass
    # 즉시 종료
    os._exit(1)

# 시그널 핸들러 등록
signal.signal(signal.SIGINT, signal_handler)

def check_interrupted():
    """중단 신호를 확인하고 중단된 경우 예외를 발생시킵니다."""
    global interrupted
    if interrupted:
        raise KeyboardInterrupt("사용자에 의해 중단되었습니다.")

def setup_driver(headless=False):
    """Selenium WebDriver를 설정하고 다운로드 폴더를 지정합니다."""
    options = webdriver.ChromeOptions()
    # 다운로드 폴더 설정
    if not os.path.exists(config.DOWNLOAD_FOLDER):
        os.makedirs(config.DOWNLOAD_FOLDER)
    prefs = {"download.default_directory": config.DOWNLOAD_FOLDER}
    options.add_experimental_option("prefs", prefs)
    
    # 창 크기 및 안정성 옵션 추가
    if headless:
        options.add_argument('--headless')
        options.add_argument('--window-size=1920,1080')
    else:
        options.add_argument("--start-maximized")  # 창을 최대화하여 시작
        
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--disable-extensions')
    options.add_argument('--disable-logging')
    options.add_argument('--disable-web-security')
    options.add_argument('--allow-running-insecure-content')
    options.add_argument('--disable-features=VizDisplayCompositor')
    options.add_argument('--remote-debugging-port=9222')
    options.add_experimental_option("useAutomationExtension", False)
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    
    # ChromeDriver 서비스 설정 개선
    try:
        # ChromeDriver 경로를 명시적으로 지정
        chromedriver_path = ChromeDriverManager().install()
        # print(f"ChromeDriver 경로: {chromedriver_path}")
        
        service = Service(chromedriver_path)
        driver = webdriver.Chrome(service=service, options=options)
        
    except Exception as e:
        print(f"ChromeDriver 생성 실패: {e}")
        print("Chrome 브라우저를 수동으로 열고 비플로우에 로그인한 후 다시 시도해주세요.")
        print("또는 Chrome 브라우저를 완전히 종료한 후 다시 시도해주세요.")
        raise e
    
    # 페이지 로드 타임아웃 설정
    driver.set_page_load_timeout(60)
    # 암시적 대기 시간을 20초로 늘려 안정성 확보
    driver.implicitly_wait(20)
    return driver

def authenticate_google_sheets():
    """Google Sheets API 인증을 처리합니다."""
    print("Google Sheets 인증을 시작합니다...")
    try:
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file(config.GSPREAD_CREDENTIALS_PATH, scopes=scope)
        client = gspread.authorize(creds)
        print("✅ Google Sheets 인증 성공!")
        print(f"📊 접근할 스프레드시트: '{config.SPREADSHEET_NAME}'")
        print(f"📋 접근할 시트: '{config.SOURCE_SHEET_NAME}'")
        return client
    except FileNotFoundError:
        print(f"❌ 오류: 서비스 계정 키 파일을 찾을 수 없습니다. config.GSPREAD_CREDENTIALS_PATH 경로를 확인하세요.")
    except Exception as e:
        print(f"❌ Google Sheets 인증 중 오류 발생: {e}")
    return None

def rename_downloaded_file(download_folder):
    """다운로드된 파일을 지정된 형식으로 이름을 변경합니다."""
    try:
        # 다운로드 폴더의 모든 파일 확인
        files = os.listdir(download_folder)
        if not files:
            return None
            
        # 가장 최근에 생성된 파일 찾기 (엑셀 파일 우선)
        excel_files = [f for f in files if f.endswith(('.xlsx', '.xls'))]
        if excel_files:
            latest_file = max(excel_files, key=lambda x: os.path.getctime(os.path.join(download_folder, x)))
        else:
            latest_file = max(files, key=lambda x: os.path.getctime(os.path.join(download_folder, x)))
        
        # 새 파일명 생성 (시트 탭명 사용)
        new_filename = f"{config.SOURCE_SHEET_NAME}.xlsx"
        new_filepath = os.path.join(download_folder, new_filename)
        old_filepath = os.path.join(download_folder, latest_file)
        
        # 파일명 변경
        if os.path.exists(old_filepath):
            shutil.move(old_filepath, new_filepath)
            print(f"✅ 파일명 변경 완료: {latest_file} → {new_filename}")
            return new_filepath
        else:
            print(f"⚠️ 파일을 찾을 수 없습니다: {old_filepath}")
            return None
            
    except Exception as e:
        print(f"❌ 파일명 변경 중 오류 발생: {e}")
        return None

def find_column_by_header(sheet, header_texts, start_row=1):
    """헤더 텍스트로 열 번호를 찾습니다."""
    try:
        header_row = sheet[start_row]
        for col_idx, cell in enumerate(header_row, 1):
            cell_value = str(cell.value or '').strip()
            for header_text in header_texts:
                if header_text.lower() in cell_value.lower():
                    return col_idx
        return None
    except Exception as e:
        print(f"⚠️ 헤더 검색 중 오류: {e}")
        return None

def get_data_from_sheet(client):
    """
    구글 시트에서 상품번호 목록을 가져옵니다.
    """
    print(f"📊 스프레드시트명: '{config.SPREADSHEET_NAME}'")
    print(f"📋 시트명: '{config.SOURCE_SHEET_NAME}'")
    print(f"'{config.SPREADSHEET_NAME}' 스프레드시트의 '{config.SOURCE_SHEET_NAME}' 시트에서 데이터를 가져옵니다...")
    try:
        sheet = client.open(config.SPREADSHEET_NAME).worksheet(config.SOURCE_SHEET_NAME)
        all_values = sheet.get_all_values()
        
        if len(all_values) < config.START_ROW:
            print("⚠️ 시트에 데이터가 없습니다.")
            return []

        # 헤더와 데이터 분리
        data_rows = all_values[config.START_ROW - 1:]

        # 처리할 상품번호와 행 번호를 저장할 리스트
        products_to_process = []
        total_count = 0
        
        for i, row in enumerate(data_rows):
            row_num = config.START_ROW + i
            
            # 상품번호만 가져오기 (열 인덱스 벗어남 방지)
            product_id = row[config.PRODUCT_ID_COLUMN - 1].strip() if len(row) >= config.PRODUCT_ID_COLUMN else ""

            if product_id:  # 상품번호가 있는 경우에만 처리
                total_count += 1
                products_to_process.append({'product_id': product_id, 'row_num': row_num})
        
        if not products_to_process:
            print("✅ 처리할 상품이 없습니다.")
            return []

        print(f"총 {total_count}개 상품을 처리합니다.")
        return products_to_process

    except gspread.exceptions.SpreadsheetNotFound:
        print(f"❌ 오류: 스프레드시트 '{config.SPREADSHEET_NAME}'을(를) 찾을 수 없습니다.")
    except gspread.exceptions.WorksheetNotFound:
        print(f"❌ 오류: '{config.SOURCE_SHEET_NAME}' 시트를 찾을 수 없습니다.")
    except Exception as e:
        print(f"❌ 시트 데이터 로딩 중 오류 발생: {e}")
    return []

def get_bc_column_data_from_sheet(client):
    """
    구글 시트에서 B열(상품번호)과 C열(상품명) 데이터를 가져옵니다.
    """
    print(f"📊 스프레드시트명: '{config.SPREADSHEET_NAME}'")
    print(f"📋 시트명: '{config.SOURCE_SHEET_NAME}'")
    print(f"'{config.SPREADSHEET_NAME}' 스프레드시트의 '{config.SOURCE_SHEET_NAME}' 시트에서 B,C열 데이터를 가져옵니다...")
    try:
        sheet = client.open(config.SPREADSHEET_NAME).worksheet(config.SOURCE_SHEET_NAME)
        all_values = sheet.get_all_values()
        
        if len(all_values) < config.START_ROW:
            print("⚠️ 시트에 데이터가 없습니다.")
            return []

        # 헤더와 데이터 분리
        data_rows = all_values[config.START_ROW - 1:]

        # B,C열 데이터와 행 번호를 저장할 리스트
        bc_column_data = []
        
        for i, row in enumerate(data_rows):
            row_num = config.START_ROW + i
            
            # B열(상품번호)과 C열(상품명) 가져오기
            product_id = row[1].strip() if len(row) >= 2 else ""  # B열은 2번째 열 (인덱스 1)
            product_name = row[2].strip() if len(row) >= 3 else ""  # C열은 3번째 열 (인덱스 2)

            if product_id and product_name:  # 둘 다 있는 경우에만 처리
                bc_column_data.append({
                    'row_num': row_num,
                    'product_id': product_id,
                    'product_name': product_name
                })
                # print(f"행 {row_num}: B열={product_id}, C열={product_name}")
        
        if not bc_column_data:
            print("⚠️ B,C열에 유효한 데이터가 없습니다.")
            return []

        print(f"✅ B,C열에서 {len(bc_column_data)}개 데이터를 찾았습니다.")
        return bc_column_data

    except Exception as e:
        print(f"❌ B,C열 데이터 로딩 중 오류 발생: {e}")
        return []

def get_bi_column_data_from_sheet(client):
    """
    구글 시트에서 BI열(URL) 데이터를 가져옵니다.
    """
    print(f"📊 스프레드시트명: '{config.SPREADSHEET_NAME}'")
    print(f"📋 시트명: '{config.SOURCE_SHEET_NAME}'")
    print(f"'{config.SPREADSHEET_NAME}' 스프레드시트의 '{config.SOURCE_SHEET_NAME}' 시트에서 BI열 데이터를 가져옵니다...")
    try:
        sheet = client.open(config.SPREADSHEET_NAME).worksheet(config.SOURCE_SHEET_NAME)
        all_values = sheet.get_all_values()
        
        if len(all_values) < config.START_ROW:
            print("⚠️ 시트에 데이터가 없습니다.")
            return []

        # 헤더와 데이터 분리
        data_rows = all_values[config.START_ROW - 1:]

        # BI열 데이터와 행 번호를 저장할 리스트
        bi_column_data = []
        
        for i, row in enumerate(data_rows):
            row_num = config.START_ROW + i
            
            # BI열(URL) 가져오기 (61번째 열, 인덱스 60)
            url = row[60].strip() if len(row) >= 61 else ""

            if url and url.startswith('http'):  # 유효한 URL인 경우에만 처리
                bi_column_data.append({
                    'row_num': row_num,
                    'url': url
                })
                # print(f"행 {row_num}: BI열={url}")
        
        if not bi_column_data:
            print("⚠️ BI열에 유효한 URL 데이터가 없습니다.")
            return []

        print(f"✅ BI열에서 {len(bi_column_data)}개 URL을 찾았습니다.")
        return bi_column_data

    except Exception as e:
        print(f"❌ BI열 데이터 로딩 중 오류 발생: {e}")
        return []

def get_bh_column_data_from_sheet(client):
    """
    구글 시트에서 BH열(URL) 데이터를 가져옵니다.
    """
    print(f"📊 스프레드시트명: '{config.SPREADSHEET_NAME}'")
    print(f"📋 시트명: '{config.SOURCE_SHEET_NAME}'")
    print(f"'{config.SPREADSHEET_NAME}' 스프레드시트의 '{config.SOURCE_SHEET_NAME}' 시트에서 BH열 데이터를 가져옵니다...")
    try:
        sheet = client.open(config.SPREADSHEET_NAME).worksheet(config.SOURCE_SHEET_NAME)
        all_values = sheet.get_all_values()
        
        if len(all_values) < config.START_ROW:
            print("⚠️ 시트에 데이터가 없습니다.")
            return []

        # 헤더와 데이터 분리
        data_rows = all_values[config.START_ROW - 1:]

        # BH열 데이터와 행 번호를 저장할 리스트
        bh_column_data = []
        
        for i, row in enumerate(data_rows):
            row_num = config.START_ROW + i
            
            # BH열(URL) 가져오기 (60번째 열, 인덱스 59)
            url = row[59].strip() if len(row) >= 60 else ""

            if url and url.startswith('http'):  # 유효한 URL인 경우에만 처리
                bh_column_data.append({
                    'row_num': row_num,
                    'url': url
                })
                # print(f"행 {row_num}: BH열={url}")
        
        if not bh_column_data:
            print("⚠️ BH열에 유효한 URL 데이터가 없습니다.")
            return []

        print(f"✅ BH열에서 {len(bh_column_data)}개 URL을 찾았습니다.")
        return bh_column_data

    except Exception as e:
        print(f"❌ BH열 데이터 로딩 중 오류 발생: {e}")
        return []

def read_excel_data_by_product_ids(excel_file_path, product_data_list):
    """엑셀 파일에서 상품번호로 데이터를 찾아 읽어옵니다."""
    try:
        print(f"📖 엑셀 파일을 읽는 중: {excel_file_path}")
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook.active
        
        print(f"📊 엑셀 파일 정보: {sheet.max_row}행, {sheet.max_column}열")
        
        # 헤더에서 열 위치 동적 탐지
        product_id_col = find_column_by_header(sheet, ['상품번호', 'product_id', 'productid'], 1)
        product_name_col = find_column_by_header(sheet, ['상품명', 'product_name', 'productname'], 1)
        price_col = find_column_by_header(sheet, ['판매가', 'price', 'selling_price'], 1)
        option_col = find_column_by_header(sheet, ['옵션값'], 1)
        
        # 기본값 설정 (탐지 실패시)
        if not product_id_col:
            product_id_col = 2  # B열
            print("⚠️ 상품번호 열을 찾을 수 없어 B열(2)을 사용합니다.")
        if not product_name_col:
            product_name_col = 3  # C열
            print("⚠️ 상품명 열을 찾을 수 없어 C열(3)을 사용합니다.")
        if not price_col:
            price_col = 4  # D열
            print("⚠️ 판매가 열을 찾을 수 없어 D열(4)을 사용합니다.")
        if not option_col:
            option_col = 50  # AX열
            print("⚠️ 옵션 열을 찾을 수 없어 AX열(50)을 사용합니다.")
        
        # 데이터 딕셔너리 생성 (상품번호를 키로)
        matched_data = {}
        unmatched_products = []
        
        # 엑셀 데이터를 먼저 모두 읽어서 딕셔너리로 저장
        excel_data_dict = {}
        for row in range(2, sheet.max_row + 1):
            try:
                excel_product_id = sheet.cell(row=row, column=product_id_col).value
                if excel_product_id:
                    excel_product_id = str(excel_product_id).strip()
                    excel_product_name = sheet.cell(row=row, column=product_name_col).value
                    excel_price = sheet.cell(row=row, column=price_col).value
                    excel_option = sheet.cell(row=row, column=option_col).value
                    
                    excel_data_dict[excel_product_id] = {
                        'name': str(excel_product_name) if excel_product_name else '',
                        'price': str(excel_price) if excel_price else '',
                        'option': str(excel_option) if excel_option else '',
                        'row': row
                    }
            except Exception as e:
                continue
        
        print(f"📊 엑셀에서 {len(excel_data_dict)}개 상품 데이터를 로드했습니다.")
        
        # 구글 시트의 상품번호와 매칭
        for product_info in product_data_list:
            product_id = str(product_info['product_id']).strip()
            product_name = product_info['product_name']
            
            if product_id in excel_data_dict:
                excel_data = excel_data_dict[product_id]
                matched_data[product_id] = {
                    'name': excel_data['name'],
                    'price': excel_data['price'],
                    'option': excel_data['option'],
                    'google_name': product_name  # 구글 시트의 상품명도 저장
                }
            else:
                unmatched_products.append({
                    'product_id': product_id,
                    'product_name': product_name
                })
        
        # 매칭 실패한 상품들에 대해 상품명으로 재시도 (폴백)
        if unmatched_products:
            print(f"\n🔄 {len(unmatched_products)}개 상품에 대해 상품명으로 재시도합니다...")
            for product_info in unmatched_products[:]:  # 복사본으로 순회
                product_id = product_info['product_id']
                product_name = product_info['product_name']
                
                # 상품명으로 매칭 시도
                for excel_id, excel_data in excel_data_dict.items():
                    if excel_data['name'] and str(excel_data['name']).strip() == product_name:
                        matched_data[product_id] = {
                            'name': excel_data['name'],
                            'price': excel_data['price'],
                            'option': excel_data['option'],
                            'google_name': product_name
                        }
                        unmatched_products.remove(product_info)
                        break
        
        print(f"\n📊 매칭 결과: 성공 {len(matched_data)}개, 실패 {len(unmatched_products)}개")
        
        return matched_data
        
    except Exception as e:
        print(f"❌ 엑셀 파일 읽기 중 오류 발생: {e}")
        return {}

def update_google_sheet_with_excel_data(client, excel_data, bc_column_data):
    """엑셀 데이터를 구글 시트에 업데이트합니다."""
    if not excel_data or not bc_column_data:
        return
        
    print(f"\n📊 스프레드시트명: '{config.SPREADSHEET_NAME}'")
    print(f"📋 시트명: '{config.SOURCE_SHEET_NAME}'")
    print(f"📝 구글 시트에 엑셀 데이터를 업데이트합니다...")
    try:
        sheet = client.open(config.SPREADSHEET_NAME).worksheet(config.SOURCE_SHEET_NAME)
        
        cells_to_update = []
        updated_count = 0
        
        for bc_data in bc_column_data:
            row_num = bc_data['row_num']
            product_id = bc_data['product_id']
            product_name = bc_data['product_name']
            
            # 엑셀에서 해당 상품번호의 데이터 찾기
            if product_id in excel_data:
                data = excel_data[product_id]
                
                # F열(6): 상품명, H열(8): 상품가, I열(9): 옵션값
                cells_to_update.append(gspread.Cell(row_num, 6, data['name']))  # F열
                cells_to_update.append(gspread.Cell(row_num, 8, data['price']))  # H열
                cells_to_update.append(gspread.Cell(row_num, 9, data['option']))  # I열
                
                updated_count += 1
            else:
                print(f"⚠️ 행 {row_num} (상품번호: {product_id}, 상품명: '{product_name}'): 엑셀에서 해당 상품을 찾을 수 없습니다.")
        
        if cells_to_update:
            sheet.update_cells(cells_to_update, value_input_option='USER_ENTERED')
            print(f"✅ 구글 시트 업데이트 완료! ({updated_count}개 상품)")
        else:
            print("⚠️ 업데이트할 데이터가 없습니다.")
            
    except Exception as e:
        print(f"❌ 구글 시트 업데이트 중 오류 발생: {e}")

def update_sheet_status(client, processed_rows):
    """
    처리가 완료된 행에 대해 구글 시트의 상태를 업데이트합니다.
    """
    if not processed_rows:
        return
    
    print(f"Google 시트에 {len(processed_rows)}개 항목의 작업 상태를 업데이트합니다...")
    try:
        sheet = client.open(config.SPREADSHEET_NAME).worksheet(config.SOURCE_SHEET_NAME)
        
        cells_to_update = []
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        for row_num in processed_rows:
            # D열(4): 작업 상태만 업데이트 (E열은 건드리지 않음)
            cells_to_update.append(gspread.Cell(row_num, 4, '다운로드 완료'))  # D열

        sheet.update_cells(cells_to_update, value_input_option='USER_ENTERED')
        print("✅ Google 시트 상태 업데이트 완료!")

    except Exception as e:
        print(f"❌ 시트 업데이트 중 오류 발생: {e}")

def search_and_download_naver_format(driver, products_to_process, client=None):
    """
    b-flow에서 상품을 검색하고 '네이버 스마트스토어 형식'으로 다운로드합니다.
    성공적으로 처리된 chunk의 행 번호 목록을 yield합니다.
    """
    # 로그인
    print(f"'{config.SEARCH_SITE_URL}'에 접속합니다.")
    driver.get(config.SEARCH_SITE_URL)
    
    # --- 웹사이트 로그인 로직 ---
    try:
        print("로그인을 시작합니다...")
        login_button = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, "/html/body/div[1]/div[3]/div[1]/div[2]/button[2]"))
        )
        login_button.click()
        
        username_input = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((By.XPATH, "/html/body/div[1]/div[14]/div/div[2]/div/div[2]/div/input[1]"))
        )
        username_input.send_keys(config.BFLOW_ID)
        
        password_input = driver.find_element(By.XPATH, "/html/body/div[1]/div[14]/div/div[2]/div/div[2]/div/input[2]")
        password_input.send_keys(config.BFLOW_PW)
        
        submit_button = driver.find_element(By.XPATH, "/html/body/div[1]/div[14]/div/div[2]/div/div[3]/button[1]")
        submit_button.click()
        
        print("로그인 성공. 페이지 로딩을 기다립니다...")
        WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#main-page"))
        )
        print("메인 페이지 로딩 완료.")
    except Exception as e:
        print(f"로그인 중 오류가 발생했습니다: {e}")
        return []
    # --- 로그인 로직 끝 ---

    # 상품조회/수정 페이지로 이동
    print("상품 조회/수정 페이지로 이동합니다...")
    driver.get("https://b-flow.co.kr/products/new#/")
    try:
        WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.box.collapsed-box")))
        print("✅ 상품 조회 페이지 로딩 완료.")
    except Exception as e:
        print(f"❌ 페이지 로딩 대기 중 오류: {e}")
        return

    chunks = [products_to_process[i:i + 500] for i in range(0, len(products_to_process), 500)]
    
    for i, chunk in enumerate(chunks):
        product_ids_in_chunk = [p['product_id'] for p in chunk]
        row_nums_in_chunk = [p['row_num'] for p in chunk]

        for attempt in range(config.RETRY_COUNT):
            try:
                print(f"\n--- 묶음 {i+1}/{len(chunks)} 처리 시작 (상품 {len(product_ids_in_chunk)}개), 시도 {attempt + 1}/{config.RETRY_COUNT} ---")
                
                # 검색 필터 '상품번호'로 설정
                search_filter_dropdown = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.multiselect.br-select")))
                search_filter_dropdown.click()
                product_number_option = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//ul[contains(@class, 'multiselect__content')]//span[contains(text(), '상품번호')]")))
                product_number_option.click()
                
                # 상품번호 입력
                search_box = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div.br-text-wrapper > input")))
                search_box.clear()
                search_box.send_keys(" \n".join(product_ids_in_chunk))
                time.sleep(1)

                # 검색 버튼 클릭
                driver.find_element(By.CSS_SELECTOR, "button.br-btn-purple").click()
                print("🔍 검색 실행... 결과를 기다립니다.")

                excel_dropdown = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, "//button[contains(., '엑셀 다운로드')]")))
                print("✅ 검색 결과 로딩 완료.")
                
                # 검색 결과 완전 로딩을 위한 추가 대기
                print("⏳ 검색 결과 테이블 로딩을 기다립니다...")
                time.sleep(3)
                
                # 테이블이 완전히 로딩될 때까지 대기
                WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr")))
                print("✅ 테이블 데이터 로딩 완료.")

                # 전체선택 체크박스 클릭 (커스텀 체크박스의 부모 div 클릭)
                print("📋 전체선택 체크박스를 클릭합니다.")
                time.sleep(2)  # 체크박스 클릭 전 추가 대기
                try:
                    # CSS 셀렉터로 시도
                    select_all_checkbox = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "#main-page > div > div > section > div > div:nth-child(4) > div.box-body > div > table > thead > tr:nth-child(2) > th:nth-child(1) > div")))
                    select_all_checkbox.click()
                    time.sleep(1)  # 클릭 후 대기
                except:
                    # XPath로 대체 시도
                    select_all_checkbox = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, "//table//thead//tr[2]//th[1]//div[@p-checkbox]")))
                    select_all_checkbox.click()
                    time.sleep(1)  # 클릭 후 대기
                print("✅ 전체선택 완료.")

                initial_file_count = len(os.listdir(config.DOWNLOAD_FOLDER))
                
                # 엑셀 다운로드 버튼 클릭 전 대기
                print("📊 엑셀 다운로드 버튼을 클릭합니다.")
                time.sleep(1)
                excel_dropdown.click()
                time.sleep(2)  # 드롭다운 메뉴 로딩 대기
                
                # --- [핵심] "네이버 스마트스토어 형식" 클릭 ---
                print("📂 '네이버 스마트스토어 형식' 다운로드를 선택합니다.")
                download_option = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.XPATH, "//ul[contains(@class, 'dropdown-menu')]//a[normalize-space()='네이버 스마트스토어 형식']")))
                download_option.click()
                time.sleep(2)  # 다운로드 옵션 선택 후 대기
                
                print("⚠️ 알림창을 확인합니다...")
                WebDriverWait(driver, 15).until(EC.alert_is_present()).accept()
                print("✅ 알림창 확인 완료.")

                print("⏳ 다운로드 모달 로딩을 기다립니다...")
                WebDriverWait(driver, 60).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "div.v--modal-box span.br-label-green")))
                time.sleep(2)  # 모달 완전 로딩 대기
                
                print("🚀 파일 다운로드 버튼을 클릭합니다.")
                driver.find_element(By.CSS_SELECTOR, "div.v--modal-box tbody > tr:nth-child(1) > td:nth-child(7) > button").click()
                print("🚀 파일 다운로드를 시작합니다.")

                # 다운로드 완료 대기 (더 긴 시간으로 확장)
                print("⏳ 파일 다운로드 완료를 기다립니다...")
                download_completed = False
                downloaded_file_path = None
                for i in range(60):  # 60초로 확장
                    if len(os.listdir(config.DOWNLOAD_FOLDER)) > initial_file_count:
                        print("✅ 파일 다운로드 완료.")
                        # 다운로드된 파일명 변경
                        downloaded_file_path = rename_downloaded_file(config.DOWNLOAD_FOLDER)
                        download_completed = True
                        break
                    if i % 10 == 0 and i > 0:  # 10초마다 진행 상황 출력
                        print(f"⏳ 다운로드 대기 중... ({i}/60초)")
                    time.sleep(1)
                
                if not download_completed:
                    print("⚠️ 파일 다운로드를 감지하지 못했습니다.")
                
                # 모달 닫기 전 대기
                print("📋 다운로드 모달을 닫습니다.")
                time.sleep(2)
                driver.find_element(By.CSS_SELECTOR, "div.v--modal-box span.close-btn").click()
                time.sleep(3)  # 모달 닫기 후 충분한 대기
                
                # 엑셀 데이터 읽기 및 구글 시트 업데이트
                if downloaded_file_path and os.path.exists(downloaded_file_path):
                    print(f"\n📊 엑셀 데이터를 구글 시트에 업데이트합니다...")
                    
                    # B,C열 데이터 가져오기
                    bc_data = get_bc_column_data_from_sheet(client)
                    if bc_data:
                        # B,C열 데이터를 사용하여 엑셀 데이터 읽기 (상품번호 기준)
                        excel_data = read_excel_data_by_product_ids(downloaded_file_path, bc_data)
                        
                        if excel_data:
                            print(f"✅ 엑셀에서 {len(excel_data)}개 상품 데이터를 읽었습니다.")
                            # 구글 시트 업데이트
                            update_google_sheet_with_excel_data(client, excel_data, bc_data)
                            print("✅ 엑셀 데이터로 구글 시트 업데이트 완료!")
                        else:
                            print("❌ 엑셀 데이터를 읽을 수 없습니다.")
                            excel_data = {}
                    else:
                        print("❌ B,C열 데이터를 가져올 수 없습니다.")
                        excel_data = {}
                    
                    # 현재 처리된 상품들에 대한 정보를 yield로 전달
                    yield {'row_nums': row_nums_in_chunk, 'excel_data': excel_data, 'products': chunk}
                else:
                    yield {'row_nums': row_nums_in_chunk, 'excel_data': {}, 'products': chunk}
                
                print(f"--- 묶음 {i+1} 처리 성공 ---")
                
                # 다음 chunk를 위해 페이지 새로고침
                print("🔄 다음 배치를 위해 페이지를 새로고침합니다.")
                driver.refresh()
                time.sleep(3)  # 새로고침 후 대기
                WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.box.collapsed-box")))
                print("✅ 페이지 새로고침 완료.")
                break
            
            except Exception as e:
                print(f"❌ 오류 발생 (시도 {attempt + 1}/{config.RETRY_COUNT}): {e}")
                driver.refresh()
                WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div.box.collapsed-box")))
                time.sleep(3)
                if attempt == config.RETRY_COUNT - 1:
                    print(f"❌ 최대 재시도 횟수({config.RETRY_COUNT})를 초과하여 이번 묶음을 건너뜁니다.")

def extract_image_url_from_brich(driver, url):
    """
    Selenium을 사용하여 brich.co.kr URL에서 상품 이미지 URL을 추출합니다.
    """
    global interrupted
    
    # 중단 신호 확인
    if interrupted:
        return None
        
    try:
        # 타임아웃 설정 (중요!)
        driver.set_page_load_timeout(30)  # 페이지 로드 타임아웃
        driver.implicitly_wait(5)  # 요소 찾기 타임아웃
        
        # 페이지 로드
        driver.get(url)
        
        # 활성화된 swiper-slide의 product-image-swipe div 찾기
        try:
            # 먼저 활성화된 슬라이드의 이미지를 찾기 (타임아웃 5초)
            active_image_div = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "div.swiper-slide-active div.product-image-swipe"))
            )
            
            # style 속성에서 background-image URL 추출
            style = active_image_div.get_attribute('style')
            
            if style and 'background-image' in style:
                # url("...") 패턴에서 URL 추출
                import re
                url_match = re.search(r'url\(["\']?([^"\']+)["\']?\)', style)
                if url_match:
                    image_url = url_match.group(1)
                    return image_url
            
        except Exception as e:
            # 첫 번째 product-image-swipe div 찾기 (fallback)
            try:
                first_image_div = driver.find_element(By.CSS_SELECTOR, "div.product-image-swipe")
                style = first_image_div.get_attribute('style')
                
                if style and 'background-image' in style:
                    import re
                    url_match = re.search(r'url\(["\']?([^"\']+)["\']?\)', style)
                    if url_match:
                        image_url = url_match.group(1)
                        return image_url
                        
            except Exception as e2:
                # 모든 product-image-swipe div 찾기
                try:
                    all_image_divs = driver.find_elements(By.CSS_SELECTOR, "div.product-image-swipe")
                    
                    for i, div in enumerate(all_image_divs):
                        style = div.get_attribute('style')
                        
                        if style and 'background-image' in style:
                            import re
                            url_match = re.search(r'url\(["\']?([^"\']+)["\']?\)', style)
                            if url_match:
                                image_url = url_match.group(1)
                                return image_url
                                
                except Exception as e3:
                    pass
        
        print(f"⚠️ 이미지 URL을 찾을 수 없습니다: {url}")
        return None
        
    except Exception as e:
        print(f"❌ 이미지 URL 추출 중 오류 발생 ({url}): {e}")
        return None

def update_bj_column_with_image_urls(client, url_column_data, column_type="BI"):
    """
    BI열 또는 BH열의 URL들을 사용하여 BJ열에 이미지 URL을 업데이트합니다.
    병렬 처리를 사용하여 속도를 개선합니다.
    """
    global interrupted
    
    if not url_column_data:
        return
    print(f"📋 시트명: '{config.SOURCE_SHEET_NAME}'")
    print(f"📝 {column_type}열의 URL들을 사용하여 BJ열에 이미지 URL을 업데이트합니다...")
    print(f"🚀 병렬 처리 시작 (워커 수: {config.MAX_WORKERS})")
    
    # 데이터를 워커 수에 맞게 분할
    chunk_size = (len(url_column_data) + config.MAX_WORKERS - 1) // config.MAX_WORKERS
    chunks = [url_column_data[i:i + chunk_size] for i in range(0, len(url_column_data), chunk_size)]
    
    all_results = []
    
    try:
        with ThreadPoolExecutor(max_workers=config.MAX_WORKERS) as executor:
            futures = [executor.submit(process_url_chunk, chunk, i+1) for i, chunk in enumerate(chunks)]
            
            for future in as_completed(futures):
                if interrupted:
                    executor.shutdown(wait=False, cancel_futures=True)
                    break
                try:
                    result = future.result()
                    all_results.extend(result)
                except Exception as e:
                    print(f"❌ 병렬 처리 중 오류 발생: {e}")
                    
    except KeyboardInterrupt:
        print("\n⚠️ 사용자에 의해 중단되었습니다.")
        return

    # 결과 취합 및 구글 시트 업데이트
    if all_results:
        print(f"✅ 총 {len(all_results)}개의 이미지 URL을 추출했습니다. 구글 시트에 업데이트합니다...")
        
        try:
            sheet = client.open(config.SPREADSHEET_NAME).worksheet(config.SOURCE_SHEET_NAME)
            cells_to_update = []
            
            for item in all_results:
                row_num = item['row_num']
                image_url = item['image_url']
                # BJ열(62번째 열)에 이미지 URL 업데이트
                cells_to_update.append(gspread.Cell(row_num, 62, image_url))
            
            if cells_to_update:
                sheet.update_cells(cells_to_update, value_input_option='USER_ENTERED')
                print(f"✅ BJ열 업데이트 완료! ({len(cells_to_update)}개 이미지 URL)")
            else:
                print("⚠️ 업데이트할 이미지 URL이 없습니다.")
                
        except Exception as e:
            print(f"❌ BJ열 업데이트 중 오류 발생: {e}")
    else:
        print("⚠️ 추출된 이미지 URL이 없습니다.")

def update_image_urls_only():
    """BI열 또는 BH열의 URL들로부터 BJ열에 이미지 URL만 업데이트합니다."""
    print("🖼️ 이미지 URL 추출 및 BJ열 업데이트를 시작합니다...")
    
    # 구글 시트 인증
    gspread_client = authenticate_google_sheets()
    if not gspread_client:
        return

    # BI열과 BH열 데이터 모두 확인
    bi_column_data = get_bi_column_data_from_sheet(gspread_client)
    bh_column_data = get_bh_column_data_from_sheet(gspread_client)
    
    # 어떤 열에 데이터가 있는지 확인
    if bi_column_data and bh_column_data:
        print(f"📋 BI열에서 {len(bi_column_data)}개, BH열에서 {len(bh_column_data)}개 URL을 찾았습니다.")
        print("⚠️ 두 열 모두에 데이터가 있습니다. BI열을 우선 처리합니다.")
        url_data = bi_column_data
        column_type = "BI"
    elif bi_column_data:
        print(f"📋 BI열에서 {len(bi_column_data)}개 URL을 찾았습니다.")
        url_data = bi_column_data
        column_type = "BI"
    elif bh_column_data:
        print(f"📋 BH열에서 {len(bh_column_data)}개 URL을 찾았습니다.")
        url_data = bh_column_data
        column_type = "BH"
    else:
        print("⚠️ BI열과 BH열 모두에 URL 데이터가 없습니다.")
        return
    
    # BJ열에 이미지 URL 업데이트
    update_bj_column_with_image_urls(gspread_client, url_data, column_type)
    
    print("\n🎉 이미지 URL 업데이트가 완료되었습니다!")

def update_sheet_from_existing_file():
    """기존 엑셀 파일을 사용하여 구글 시트를 업데이트합니다."""
    print("📊 기존 엑셀 파일로 구글 시트를 업데이트합니다...")
    
    # 구글 시트 인증
    gspread_client = authenticate_google_sheets()
    if not gspread_client:
        return

    # B,C열 데이터 가져오기
    bc_column_data = get_bc_column_data_from_sheet(gspread_client)
    if not bc_column_data:
        print("⚠️ B,C열에 유효한 데이터가 없습니다.")
        return

    # 엑셀 파일 경로 (시트 탭명 사용)
    excel_file_path = os.path.join(config.DOWNLOAD_FOLDER, f"{config.SOURCE_SHEET_NAME}.xlsx")
    
    if not os.path.exists(excel_file_path):
        print(f"❌ 엑셀 파일을 찾을 수 없습니다: {excel_file_path}")
        return

    # B,C열 데이터를 사용하여 엑셀 데이터 읽기 (상품번호 기준)
    excel_data = read_excel_data_by_product_ids(excel_file_path, bc_column_data)
    if not excel_data:
        print("❌ 엑셀 데이터를 읽을 수 없습니다.")
        return

    # 구글 시트 업데이트 (F, H, I열)
    update_google_sheet_with_excel_data(gspread_client, excel_data, bc_column_data)
    
    # BI열과 BH열 데이터 모두 확인
    bi_column_data = get_bi_column_data_from_sheet(gspread_client)
    bh_column_data = get_bh_column_data_from_sheet(gspread_client)
    
    # 어떤 열에 데이터가 있는지 확인
    if bi_column_data and bh_column_data:
        print(f"📋 BI열에서 {len(bi_column_data)}개, BH열에서 {len(bh_column_data)}개 URL을 찾았습니다.")
        print("⚠️ 두 열 모두에 데이터가 있습니다. BI열을 우선 처리합니다.")
        url_data = bi_column_data
        column_type = "BI"
    elif bi_column_data:
        print(f"📋 BI열에서 {len(bi_column_data)}개 URL을 찾았습니다.")
        url_data = bi_column_data
        column_type = "BI"
    elif bh_column_data:
        print(f"📋 BH열에서 {len(bh_column_data)}개 URL을 찾았습니다.")
        url_data = bh_column_data
        column_type = "BH"
    else:
        print("⚠️ BI열과 BH열 모두에 URL 데이터가 없어 BJ열 업데이트를 건너뜁니다.")
        print("\n🎉 구글 시트 업데이트가 완료되었습니다!")
        return
    
    # BJ열에 이미지 URL 업데이트
    update_bj_column_with_image_urls(gspread_client, url_data, column_type)
    
    print("\n🎉 구글 시트 업데이트가 완료되었습니다!")

def main():
    """자동화 스크립트의 메인 실행 함수"""
    gspread_client = authenticate_google_sheets()
    if not gspread_client:
        return

    products_to_process = get_data_from_sheet(gspread_client)
    if not products_to_process:
        return

    driver = setup_driver()
    try:
        # 각 chunk 처리 직후 시트 상태를 업데이트합니다.
        for result in search_and_download_naver_format(driver, products_to_process, gspread_client):
            if isinstance(result, dict):
                # 새로운 형식: 엑셀 데이터 포함
                processed_rows = result['row_nums']
                excel_data = result['excel_data']
                products = result['products']
                
                # 구글 시트 상태 업데이트
                update_sheet_status(gspread_client, processed_rows)
                
                # 엑셀 데이터를 구글 시트에 업데이트
                if excel_data:
                    update_google_sheet_with_excel_data(gspread_client, excel_data, products)
            else:
                # 기존 형식: 행 번호만
                update_sheet_status(gspread_client, result)
        
        print("\n🎉 모든 자동화 작업이 성공적으로 완료되었습니다.")
    finally:
        print("WebDriver를 종료합니다.")
        driver.quit()

def run_full_process(skip_download=False, extract_images=True):
    """전체 프로세스 실행 (b-flow 다운로드 + 데이터 업데이트 + 이미지 추출 선택)"""
    global interrupted
    
    print("🚀 전체 프로세스를 시작합니다...")
    print("💡 중단하려면 Ctrl+C를 누르세요.")
    
    # 1단계: 구글 시트 연결
    print("\n=== 1단계: 구글 시트 연결 ===")
    if interrupted:
        print("⚠️ 사용자에 의해 중단되었습니다.")
        return
    client = authenticate_google_sheets()
    if not client:
        print("❌ 구글 시트 연결 실패")
        return
    print("✅ 구글 시트 연결 성공")
    
    # 2단계: B열 상품번호 가져오기
    print("\n=== 2단계: B열 상품번호 가져오기 ===")
    products_to_process = get_data_from_sheet(client)
    if not products_to_process:
        print("❌ B열에 상품번호가 없습니다")
        return
    print(f"✅ B열에서 {len(products_to_process)}개 상품번호 찾음")
    
    # 3단계: b-flow에서 엑셀 다운로드 (skip_download가 False일 때만)
    if not skip_download:
        print("\n=== 3단계: b-flow에서 엑셀 다운로드 ===")
        print("b-flow에 로그인하여 상품을 검색하고 엑셀을 다운로드합니다...")
        
        driver = setup_driver()
        try:
            # b-flow 로그인 및 다운로드
            for result in search_and_download_naver_format(driver, products_to_process, client):
                if isinstance(result, dict):
                    processed_rows = result['row_nums']
                    excel_data = result['excel_data']
                    products = result['products']
                    
                    # 구글 시트 상태 업데이트
                    update_sheet_status(client, processed_rows)
                else:
                    update_sheet_status(client, result)
            
            print("✅ b-flow 다운로드 및 구글 시트 업데이트 완료!")
            
        finally:
            print("WebDriver를 종료합니다.")
            driver.quit()
    else:
        print("\n=== 3단계: b-flow 다운로드 건너뜀 ===")
        print("✅ 다운로드 단계를 건너뜁니다.")
    
    # 4단계: BI열과 BH열 데이터 확인
    print("\n=== 4단계: BI열과 BH열 데이터 확인 ===")
    bi_data = get_bi_column_data_from_sheet(client)
    bh_data = get_bh_column_data_from_sheet(client)
    
    # 어떤 열에 데이터가 있는지 확인
    if bi_data and bh_data:
        print(f"✅ BI열에서 {len(bi_data)}개, BH열에서 {len(bh_data)}개 URL 찾음")
        print("⚠️ 두 열 모두에 데이터가 있습니다. BI열을 우선 처리합니다.")
        url_data = bi_data
        column_type = "BI"
    elif bi_data:
        print(f"✅ BI열에서 {len(bi_data)}개 URL 찾음")
        url_data = bi_data
        column_type = "BI"
    elif bh_data:
        print(f"✅ BH열에서 {len(bh_data)}개 URL 찾음")
        url_data = bh_data
        column_type = "BH"
    else:
        print("❌ BI열과 BH열 모두에 URL 데이터 없음")
        print("🎉 b-flow 다운로드 및 데이터 업데이트만 완료되었습니다!")
        return
    
    # 5단계: 이미지 추출 여부 선택
    if extract_images:
        print("\n=== 5단계: 이미지 URL 추출 ===")
        print(f"{column_type}열에 {len(url_data)}개 URL이 있습니다.")
        print("🖼️ 이미지 URL 추출을 시작합니다...")
        update_bj_column_with_image_urls(client, url_data, column_type)
        print("✅ 이미지 URL 업데이트 완료!")
    else:
        print("\n=== 5단계: 이미지 추출 건너뜀 ===")
        print("✅ 이미지 추출을 건너뜁니다.")
    
    print("\n🎉 전체 프로세스가 완료되었습니다!")

def run_image_extraction_only():
    """이미지 추출만 실행"""
    print("🖼️ 이미지 추출만 실행합니다...")
    run_full_process(skip_download=True, extract_images=True)

def run_download_only():
    """다운로드만 실행 (이미지 추출 제외)"""    
    print("📥 다운로드만 실행합니다...")
    run_full_process(skip_download=False, extract_images=False)

if __name__ == '__main__':
    try:
        print("🚀 전체 프로세스를 시작합니다...")
        run_full_process(skip_download=False, extract_images=True)
    except KeyboardInterrupt:
        print("\n\n⚠️ 사용자에 의해 중단되었습니다.")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ 예상치 못한 오류가 발생했습니다: {e}")
        sys.exit(1)